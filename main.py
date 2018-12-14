#!/usr/bin/env python
# Please update inventory.xlsx for IP and Credentials
import sys
import paramiko
import time
import re
import ipaddress
import json
import textfsm
from openpyxl import load_workbook
from openpyxl import Workbook
from netmiko import ConnectHandler
from datetime import datetime

def time_log():
    time_log_stamp = datetime.now()
    log_time = time_log_stamp.strftime('%Y-%m-%d %H:%M:%S')
    return log_time

def log(write, filename):
    with open(filename, 'a') as writelog:
        #output = '%s : ' % time_log()
        output = "{} : ".format(time_log())
        output += write
        writelog.write(output)

def sw_version(result, device_type):
    if device_type == 'cisco_ios':
        ios_sw_re = re.compile(r'Version (\S+),')
        ios_pn_re = re.compile(r'Model\s+number\s+:\s+(\S+)')

        ios_sw = ios_sw_re.search(result)
        ios_pn = ios_pn_re.search(result)

        if ios_sw:
            sw = ios_sw.group(1)
            if ios_pn:
                pn = ios_pn.group(1)
                return sw, pn

    elif device_type == 'cisco_xe':
        xe_sw_re = re.compile(r'Version (\S+)')
        xe_pn_re = re.compile(r'cisco (WS-\S+)')

        xe_sw = xe_sw_re.search(result)
        xe_pn = xe_pn_re.search(result)

        if xe_sw:
            sw = xe_sw.group(1)
            if xe_pn:
                pn = xe_pn.group(1)                    
                return sw, pn

    elif device_type == 'cisco_nx':
        nx_sw_re = re.compile(r'system:\s+version\s+(\S+)')
        nx_pn_re = re.compile(r'cisco (Nexus\w+ \S+)')

        nx_sw = nx_sw_re.search(result)
        nx_pn = nx_pn_re.search(result)

        if nx_sw:
            sw = nx_sw.group(1)
            if nx_pn:
                pn = nx_pn.group(1)
                return sw, pn

def excel_to_lists(filename):
    wb = load_workbook(filename)
    ws = wb.active
    invent_list = list()
    for i in range(2, ws.max_row+1):
        invent_dict = {
            'hostname': None,
            'address': None,
            'type': None,
            'username': None,
            'password': None
        }

        invent_dict['hostname'] = ws.cell(row=i, column=1).value
        invent_dict['address'] = ws.cell(row=i, column=2).value
        invent_dict['type'] = ws.cell(row=i, column=3).value
        invent_dict['username'] = ws.cell(row=i, column=4).value
        invent_dict['password'] = ws.cell(row=i, column=5).value
        invent_list.append(invent_dict)
    return invent_list


def main():
    #inventory_file = 'data/inventory.xlsx'
    data_log = 'syslog/log.txt'
    print('-'*80)
    #device_list = excel_to_lists(inventory_file)
    while True:
        #print("\n")
        print('-------------------------- CISCO Network Tools --------------------------------')
        print('--- Please make sure you have filled out inventory data in "data" directory ---')
        print('-'*80)
        print('[1] Collect Software Version')
        print('[2] Backup Configuration')
        print('[3] Collect VLAN Database')
        print('[4] Set Common Config')
        print('[q] exit\n')
        input_select = input('Please select function above : ')
        input_select = str(input_select)

        if input_select == 'q' or input_select == 'Q':
            sys.exit()
        elif input_select == '1':
            inventory_file = 'data/inventory.xlsx'
            template = open('template/cisco_ios_show_version.template')
            print('-'*100)
            device_list = excel_to_lists(inventory_file)
            wb = Workbook()
            ws = wb.active
            swinvent_dir = 'data/software_version.xlsx'
            ws.cell(row=1, column=1, value='Hostname')
            ws.cell(row=1, column=2, value='IP Address')
            ws.cell(row=1, column=3, value='Part Number')
            ws.cell(row=1, column=4, value='Software Version')
            ws.cell(row=1, column=5, value='Uptime')
            wb.save(swinvent_dir)
            row = 2
            for device in device_list:
                device_info = {
                     'device_type': device["type"],
                     'ip': device["address"],
                     'username': device["username"],
                     'password': device["password"],
                }

                netconnect = ConnectHandler(**device_info)
                #get the real device name, not from Excel file
                dev_hostname = netconnect.find_prompt()[:-1]
                result = netconnect.send_command("show version")
                netconnect.disconnect()
                #parse output 
                sw_inventory = sw_version(result, device["type"])
                result_template = textfsm.TextFSM(template)
                version_data = result_template.ParseText(result)
                #print(version_data[0][3])
                uptime = str(version_data[0][3])
                #
                print('{} --> partnumber : {}, sw version : {}, uptime : {}'.format(device['hostname'], sw_inventory[1], sw_inventory[0], uptime))
                log('{} --> {} Uptime : {} \n'.format(device['hostname'], sw_inventory, uptime), data_log)
                #print('{} --> partnumber : {}, sw version : {}'.format(device['hostname'], sw_inventory[1], sw_inventory[0]))
                #log('{} --> {}\n'.format(device['hostname'], sw_inventory), data_log)
                #print('%s : %s --> partnumber : %s, sw version : %s' %(time_log(), device['hostname'], sw_inventory[1], sw_inventory[0]))
                #log('%s : %s --> %s\n' %(time_log(), device['hostname'], sw_inventory), data_log)
                print('-'*100)
                wb = load_workbook(swinvent_dir)
                ws = wb.active
                # Populate excel file with hostname from excel file or real hostname
                #ws.cell(row=row, column=1, value=device['hostname'])
                ws.cell(row=row, column=1, value=dev_hostname)
                ws.cell(row=row, column=2, value=device['address'])
                ws.cell(row=row, column=3, value=sw_inventory[1])
                ws.cell(row=row, column=4, value=sw_inventory[0])
                ws.cell(row=row, column=5, value=uptime)
                wb.save(swinvent_dir)
                row += 1

        elif input_select == '2':
            backup_path = "backup/{}.cfg"
            success = 0
            failure = 0
            iteration = 0
            #inventory_file = input("Please input the path of XLSX file (Ex: D:\inventory.xlsx): ")
            inventory_file = 'data/inventory.xlsx'
            print('-'*100)
            device_list = excel_to_lists(inventory_file)
            #print(json.dumps(output, indent=4))

            ssh_client = paramiko.SSHClient()
            ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            for device in device_list:
                iteration += 1
                try:
                    ipaddress.ip_address(device["address"])
                except ValueError:
                    print("Invalid IP Address Format : {} --> Please check IP Address on row number {}!".format(device["address"], iteration))
                    log("Invalid IP Address Format : {} --> Please check IP Address on row number {}!\n".format(device["address"], iteration),data_log)
                    failure += 1
                    continue
                device_info = {"ip": device["address"],
                               "username": device["username"],
                               "password": device["password"],
                               }
                #print(device_info)
                try:
                    ssh_client.connect(device_info["ip"], username=device_info["username"], password=device_info["password"])
                except:
                    print("SSH Connection Error for {}. Please check your connection or credentials!".format(device_info["ip"]))
                    log("SSH Connection Error for {}. Please check your connection or credentials!\n".format(device_info["ip"]), data_log)
                    failure += 1
                    continue
                print("Successfully connect to {}".format(device_info["ip"]))
                log("Successfully connect to {}\n".format(device_info["ip"]), data_log)
                success += 1
                ssh_conn = ssh_client.invoke_shell()
                ssh_conn.send("terminal length 0\n")
                ssh_conn.send("show run")
                ssh_conn.send("\n")
                time.sleep(2)
                if ssh_conn.recv_ready():
                    output = ssh_conn.recv(65535)
                # Python3 treat default to binary raw bits, not string implicitly --> use .decode()
                output_str = str(output.decode())
                output_str = re.search(r"!.*end", output_str, flags=re.DOTALL).group(0)
                with open(backup_path.format(device_info["ip"]), mode="w") as f:
                    f.write(output_str)
                print("Successfully backup device {}".format(device_info["ip"]))
                log("Successfully backup device {} \n".format(device_info["ip"]),data_log)
                print('-'*100)
            print("Attempt to backup {} devices".format(success+failure))
            print("Success : {}".format(success))
            print("Failure : {}".format(failure))
            log("Attempt to backup {} devices -> Success : {}, Failure : {} \n".format(success+failure,success,failure), data_log)
            print('-'*100)
        elif input_select == '3':
            inventory_file = 'data/inventory.xlsx'
            print('-'*100)
            device_list = excel_to_lists(inventory_file)
            #use TextFSM
            template = open('template/cisco_ios_show_vlan.template')
            wb = Workbook()
            ws = wb.active
            vlaninv_dir = 'data/VLAN_database.xlsx'
            wb.save(vlaninv_dir)
            for device in device_list:
                vlan_list = list()
                device_info = {
                    'device_type': device["type"],
                    'ip': device["address"],
                    'username': device["username"],
                    'password': device["password"],
                }
                netconnect = ConnectHandler(**device_info)
                dev_hostname = netconnect.find_prompt()[:-1]
                result = netconnect.send_command("show vlan")
                result_template = textfsm.TextFSM(template)
                vlan_data = result_template.ParseText(result)
                for vlan in vlan_data:
                    vlan_dict = {
                        'id': None,
                        'name': None,
                        'status': None
                    }
                    vlan_dict['id'] = vlan[0]
                    vlan_dict['name'] = vlan[1]
                    vlan_dict['status'] = vlan[2]
                    vlan_list.append(vlan_dict)
                wb = load_workbook(vlaninv_dir)
                ws = wb.create_sheet(dev_hostname)
                #wb = Workbook()
                #ws = wb.active
                ws.cell(row=1, column=1, value='VLAN ID')
                ws.cell(row=1, column=2, value='VLAN Name')
                ws.cell(row=1, column=3, value='Status')
                row = 2
                for data in vlan_list:
                    ws.cell(row=row, column=1, value=int(data['id']))
                    ws.cell(row=row, column=2, value=data['name'])
                    ws.cell(row=row, column=3, value=data['status'])
                    row += 1
                wb.save(vlaninv_dir)
                print("Collect VLAN database for {} completed".format(dev_hostname))
                log("Collect VLAN database for {} completed \n".format(dev_hostname), data_log)
                
        elif input_select == '4':
            #common_config = ["no ip http server", "no ip http secure-server"]
            config_file = "data/config.txt"
            config_f = open(config_file, "r")
            configs = config_f.read().splitlines()
            print("Setting common config :", configs)
            log('Setting common config : {}\n'.format(configs), data_log)
            #print(configs)
            inventory_file = 'data/inventory.xlsx'
            print('-'*100)
            device_list = excel_to_lists(inventory_file)
            for device in device_list:
                device_info = {
                    'device_type': device["type"],
                    'ip': device["address"],
                    'username': device["username"],
                    'password': device["password"],
                }

                netconnect = ConnectHandler(**device_info)
                print("Connected to {} ".format(device_info["ip"]))
                #print(netconnect.send_config_set(configs))
                #print(netconnect.send_command("wr mem"))
                netconnect.send_config_set(configs)
                netconnect.send_command("wr mem")
                print("Setting common config for {} completed successfully".format(device_info["ip"]))
                log('Setting common config for {} completed successfully \n'.format(device_info["ip"]), data_log)
                netconnect.disconnect()     
        else:
            print('-'*100)
            print("Please select based on the menu") 
            print('-'*100)

if __name__ == '__main__':
    main()
